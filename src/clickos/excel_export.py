"""Exporta um documento (orçamento/O.S.) para .xlsx replicando o layout impresso — editável.

Mantém o mesmo conteúdo e disposição do print (cabeçalho da empresa, dados da pessoa/veículo,
tabela de itens, resumo financeiro e condições) numa planilha sem proteção, pronta para edição.
"""
import math
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from . import markdown_min

_THIN = Side(style="thin", color="000000")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_GRAY = PatternFill("solid", fgColor="D9D9D9")
_LIGHT = PatternFill("solid", fgColor="F2F2F2")
_BOLD = Font(bold=True)
_L = Alignment(horizontal="left", vertical="center", wrap_text=True)
_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
_R = Alignment(horizontal="right", vertical="center")
COLS = "ABCDEFG"  # 7 colunas
LARGURAS = [4, 22, 13, 13, 15, 14, 14]  # A=# (estreito); demais p/ rótulos/valores largos


def _brl(v) -> str:
    try:
        n = float(v or 0)
    except (TypeError, ValueError):
        n = 0.0
    return "R$ " + f"{n:,.2f}".replace(",", "_").replace(".", ",").replace("_", ".")


def _qtd(v) -> str:
    try:
        n = float(v or 0)
    except (TypeError, ValueError):
        n = 0.0
    return f"{n:,.2f}".replace(",", "_").replace(".", ",").replace("_", ".")


def _dt(s) -> str:
    s = str(s or "")
    if len(s) >= 10 and s[4] == "-" and s[7] == "-":
        data = f"{s[8:10]}/{s[5:7]}/{s[0:4]}"
        if len(s) >= 16 and s[10] in ("T", " "):
            return f"{data} {s[11:16]}"
        return data
    return s


def gerar(doc, empresa, cliente, veiculo, prefs=None, fotos=None) -> bytes:
    empresa = empresa or {}
    cliente = cliente or {}
    veiculo = veiculo or {}
    prefs = prefs or {}
    fotos = fotos or []
    wb = Workbook()
    ws = wb.active
    ws.title = "Orcamento" if doc.get("tipo") == "orcamento" else "OrdemServico"
    ws.sheet_view.showGridLines = False
    for i, w in enumerate(LARGURAS):
        ws.column_dimensions[get_column_letter(i + 1)].width = w

    row = [1]

    def merge(r, c1, c2):
        if c1 != c2:
            ws.merge_cells(f"{c1}{r}:{c2}{r}")

    def box(r1, r2):
        for r in range(r1, r2 + 1):
            for c in COLS:
                ws[f"{c}{r}"].border = _BORDER

    def secao(texto):
        r = row[0]
        merge(r, "A", "G")
        cel = ws[f"A{r}"]
        cel.value, cel.font, cel.fill, cel.alignment = texto, Font(bold=True, size=10.5), _GRAY, _L
        box(r, r)
        row[0] += 1

    def _lbl(cell, texto):
        cell.value, cell.font, cell.fill, cell.alignment = texto, _BOLD, _LIGHT, _L

    def campo(label, valor, label2=None, valor2=None, full=False):
        """Linha rótulo/valor. Rótulo ocupa A:B (largo, sem quebrar). full=True → valor em C:G."""
        r = row[0]
        merge(r, "A", "B")
        _lbl(ws[f"A{r}"], label)
        if full:
            merge(r, "C", "G")
            ws[f"C{r}"].value, ws[f"C{r}"].alignment = (valor or ""), _L
        else:
            merge(r, "C", "D")
            ws[f"C{r}"].value, ws[f"C{r}"].alignment = (valor or ""), _L
            _lbl(ws[f"E{r}"], label2 or "")
            merge(r, "F", "G")
            ws[f"F{r}"].value, ws[f"F{r}"].alignment = (valor2 or ""), _L
        box(r, r)
        row[0] += 1

    # ---- Cabeçalho: empresa (A:E) + meta (F:G) ----
    nome_emp = empresa.get("razao_social") or empresa.get("nome_fantasia") or "Empresa"
    det = []
    if empresa.get("cnpj"):
        det.append(f"CNPJ: {empresa['cnpj']}")
    if empresa.get("ie"):
        det.append(f"I.E.: {empresa['ie']}")
    linha2 = (empresa.get("endereco") or "")
    if empresa.get("cidade"):
        linha2 += f" - {empresa['cidade']}/{empresa.get('uf') or ''}"
    if empresa.get("cep"):
        linha2 += f" - CEP: {empresa['cep']}"
    linha3 = []
    if empresa.get("telefone"):
        linha3.append(f"Tel: {empresa['telefone']}")
    if empresa.get("whatsapp"):
        linha3.append(f"WhatsApp: {empresa['whatsapp']}")
    r0 = row[0]
    merge(r0, "A", "E")
    partes = [nome_emp, "   ".join(det), linha2, "   ".join(linha3)]
    ws[f"A{r0}"].value = "\n".join(p for p in partes if p)
    ws[f"A{r0}"].font = Font(bold=True, size=12)
    ws[f"A{r0}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    merge(r0, "F", "G")
    meta = [f"Nº {doc.get('numero', '')}", f"Data: {_dt(doc.get('data_abertura'))}", f"Status: {doc.get('status', '')}"]
    if doc.get("usuario_nome"):
        meta.append(f"Atend.: {doc['usuario_nome']}")
    if doc.get("faturado_em"):
        meta.append(f"Fatur.: {_dt(doc['faturado_em'])}")
    if doc.get("ordem_compra"):
        meta.append(f"O.C.: {doc['ordem_compra']}")
    ws[f"F{r0}"].value = "\n".join(meta)
    ws[f"F{r0}"].font = Font(size=9)
    ws[f"F{r0}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[r0].height = 64
    box(r0, r0)
    row[0] += 1
    r = row[0]
    merge(r, "A", "G")
    tipo_lbl = "ORÇAMENTO" if doc.get("tipo") == "orcamento" else "ORDEM DE SERVIÇO"
    ws[f"A{r}"].value, ws[f"A{r}"].font, ws[f"A{r}"].alignment = tipo_lbl, Font(bold=True, size=13), _C
    box(r, r)
    row[0] += 1

    # ---- Dados da pessoa ----
    secao("DADOS DA PESSOA")
    campo("Razão Social / Nome", cliente.get("nome") or doc.get("cliente_nome"), "CPF / CNPJ", cliente.get("cpf_cnpj"))
    campo("Nome Fant. / Apelido", cliente.get("apelido"), "RG / IE", cliente.get("rg_ie"))
    endereco = cliente.get("endereco") or ""
    if cliente.get("numero"):
        endereco += f", {cliente['numero']}"
    if cliente.get("bairro"):
        endereco += f" - {cliente['bairro']}"
    if cliente.get("cidade"):
        endereco += f" - {cliente['cidade']}/{cliente.get('uf') or ''}"
    if cliente.get("cep"):
        endereco += f" - CEP: {cliente['cep']}"
    campo("Endereço", endereco, full=True)
    wpp_email = cliente.get("whatsapp") or ""
    if cliente.get("email"):
        wpp_email += (" · " if wpp_email else "") + cliente["email"]
    campo("Telefone", cliente.get("telefone"), "WhatsApp / E-mail", wpp_email)

    # ---- Dados do veículo ----
    secao("DADOS DO VEÍCULO")
    campo("Placa", veiculo.get("placa") or doc.get("veiculo_placa"),
          "Marca / Modelo", f"{veiculo.get('marca') or ''} {veiculo.get('modelo') or ''}".strip())
    ano = veiculo.get("ano_fab") or ""
    if veiculo.get("ano_modelo"):
        ano += f"/{veiculo['ano_modelo']}"
    campo("Versão", veiculo.get("versao"), "Ano (Fab./Mod.)", ano)
    chassi = veiculo.get("chassi") or ""
    if veiculo.get("renavam"):
        chassi += (" · " if chassi else "") + veiculo["renavam"]
    campo("Cor", veiculo.get("cor"), "Chassi / Renavam", chassi)
    campo("Combustível", veiculo.get("combustivel"), "Quilometragem", doc.get("km_entrada"))

    # ---- Checklist de entrada (somente O.S. e se o parâmetro estiver ligado) ----
    if doc.get("tipo") == "os" and str(prefs.get("os_print_checklist", "1")) == "1":
        secao("CHECKLIST DE ENTRADA DO VEÍCULO")
        lat = doc.get("lataria") or []
        metade = (len(lat) + 1) // 2
        rh = row[0]
        for col, txt in (("A", "#"), ("B", "Peça"), ("C", "Estado"), ("D", "#"), ("E", "Peça")):
            ws[f"{col}{rh}"].value, ws[f"{col}{rh}"].font, ws[f"{col}{rh}"].fill, ws[f"{col}{rh}"].alignment = txt, _BOLD, _LIGHT, _C
        merge(rh, "F", "G")
        ws[f"F{rh}"].value, ws[f"F{rh}"].font, ws[f"F{rh}"].fill, ws[f"F{rh}"].alignment = "Estado", _BOLD, _LIGHT, _C
        box(rh, rh)
        row[0] += 1
        for i in range(metade):
            e = lat[i] if i < len(lat) else {}
            d = lat[metade + i] if (metade + i) < len(lat) else None
            r = row[0]
            ws[f"A{r}"].value, ws[f"A{r}"].alignment = i + 1, _C
            ws[f"B{r}"].value, ws[f"B{r}"].alignment = (e.get("peca") or ""), _L
            ws[f"C{r}"].value, ws[f"C{r}"].alignment = (e.get("estado") or "—"), _C
            if d is not None:
                ws[f"D{r}"].value, ws[f"D{r}"].alignment = metade + i + 1, _C
                ws[f"E{r}"].value, ws[f"E{r}"].alignment = (d.get("peca") or ""), _L
                merge(r, "F", "G")
                ws[f"F{r}"].value, ws[f"F{r}"].alignment = (d.get("estado") or "—"), _C
            else:
                merge(r, "F", "G")
            box(r, r)
            row[0] += 1
        # estado geral + combustível + itens entregues
        campo("Estado Geral", doc.get("estado_geral"), "Combustível (nível)", doc.get("nivel_combustivel"))
        entregues = [n for n, k in (("Chave Principal", "item_chave_principal"), ("Chave Reserva", "item_chave_reserva"),
                                    ("Documento", "item_documento"), ("Manual", "item_manual")) if doc.get(k)]
        campo("Itens Entregues", ", ".join(entregues) or "—", full=True)
        if doc.get("ocorrencia"):
            campo("Ocorrência", doc.get("ocorrencia"), full=True)
        if doc.get("obs_entrada"):
            campo("Obs. de Entrada", doc.get("obs_entrada"), full=True)

    # ---- Itens ----
    secao("SERVIÇOS / PRODUTOS")
    rh = row[0]
    merge(rh, "B", "C")
    for col, txt in (("A", "#"), ("B", "Descrição (serviço / produto)"), ("D", "Qtd"),
                     ("E", "Vlr Bruto"), ("F", "Desconto"), ("G", "Vlr Líquido")):
        ws[f"{col}{rh}"].value, ws[f"{col}{rh}"].font, ws[f"{col}{rh}"].fill, ws[f"{col}{rh}"].alignment = txt, _BOLD, _LIGHT, _C
    box(rh, rh)
    row[0] += 1
    itens = doc.get("itens") or []
    for i, it in enumerate(itens, 1):
        r = row[0]
        merge(r, "B", "C")
        ws[f"A{r}"].value, ws[f"A{r}"].alignment = i, _C
        ws[f"B{r}"].value, ws[f"B{r}"].alignment = (it.get("descricao") or ""), _L
        for col, v in (("D", _qtd(it.get("quantidade"))), ("E", _brl(it.get("valor_unitario"))),
                       ("F", _brl(it.get("desconto"))), ("G", _brl(it.get("valor_liquido")))):
            ws[f"{col}{r}"].value, ws[f"{col}{r}"].alignment = v, _R
        box(r, r)
        row[0] += 1
    for _ in range(max(0, 4 - len(itens))):  # linhas em branco p/ preenchimento manual
        r = row[0]
        merge(r, "B", "C")
        box(r, r)
        row[0] += 1

    # ---- Resumo ----
    desc = doc.get("desconto_valor")
    if desc is None:
        desc = doc.get("desconto_geral")
    acr = doc.get("acrescimo_valor")
    if acr is None:
        acr = doc.get("acrescimo")
    for label, valor, big in (("Subtotal", doc.get("subtotal"), False), ("Desconto", desc, False),
                              ("Acréscimo", acr, False), ("TOTAL GERAL", doc.get("total"), True)):
        r = row[0]
        merge(r, "A", "E")
        f = Font(bold=True, size=12) if big else _BOLD
        ws[f"F{r}"].value, ws[f"F{r}"].font, ws[f"F{r}"].alignment = label, f, _R
        ws[f"G{r}"].value, ws[f"G{r}"].font, ws[f"G{r}"].alignment = _brl(valor), f, _R
        box(r, r)
        row[0] += 1

    # ---- Condições ----
    secao("CONDIÇÕES COMERCIAIS")
    campo("Forma de Pagamento", doc.get("forma_pagamento"), "Validade", doc.get("validade"))
    if doc.get("prazo_execucao"):
        campo("Prazo de Execução", doc.get("prazo_execucao"), full=True)
    if doc.get("observacoes"):
        campo("Observações", "", full=True)
        r = row[0] - 1
        ws[f"C{r}"].value = doc.get("observacoes")
        ws.row_dimensions[r].height = 36

    # ---- Parecer técnico (O.S.) ----
    if doc.get("parecer_mecanico") or doc.get("mecanico"):
        titulo = "PARECER TÉCNICO"
        if doc.get("mecanico"):
            titulo += f" — {doc['mecanico']}"
        secao(titulo)
        r = row[0]
        merge(r, "A", "G")
        ws[f"A{r}"].value, ws[f"A{r}"].alignment = (doc.get("parecer_mecanico") or ""), _L
        ws.row_dimensions[r].height = 38
        box(r, r)
        row[0] += 1

    # ---- Registro fotográfico (somente O.S. e se o parâmetro estiver ligado) ----
    if doc.get("tipo") == "os" and str(prefs.get("os_print_fotos", "1")) == "1" and fotos:
        from openpyxl.drawing.image import Image as XLImage
        secao("REGISTRO FOTOGRÁFICO")
        ancoras = ["A", "C", "E"]  # 3 imagens por linha
        i = 0
        while i < len(fotos):
            r = row[0]
            ws.row_dimensions[r].height = 88
            for j, col in enumerate(ancoras):
                if i + j < len(fotos):
                    try:
                        im = XLImage(BytesIO(fotos[i + j]))
                        im.width, im.height = 150, 112
                        ws.add_image(im, f"{col}{r}")
                    except Exception:
                        pass
            for c in COLS:
                ws[f"{c}{r}"].border = _BORDER
            i += 3
            row[0] += 1

    # ---- Termo de garantia (somente O.S. e se o parâmetro estiver ligado) ----
    if doc.get("tipo") == "os" and str(prefs.get("os_print_garantia", "0")) == "1" and empresa.get("termo_garantia"):
        secao("TERMO DE GARANTIA")
        blocos = markdown_min.to_blocks(empresa.get("termo_garantia")) or [{"texto": "", "negrito": False}]
        n = len(blocos)
        for idx, b in enumerate(blocos):  # uma linha por bloco; títulos em negrito; caixa ao redor
            r = row[0]
            merge(r, "A", "G")
            cel = ws[f"A{r}"]
            cel.value = b["texto"]
            cel.font = Font(size=8, bold=b["negrito"])
            cel.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            visuais = max(1, math.ceil(len(b["texto"]) / 115))  # ~115 caracteres por linha em A:G
            ws.row_dimensions[r].height = max(13, visuais * 11)
            for c in COLS:
                ws[f"{c}{r}"].border = Border(
                    top=_THIN if idx == 0 else None, bottom=_THIN if idx == n - 1 else None,
                    left=_THIN if c == "A" else None, right=_THIN if c == "G" else None)
            row[0] += 1

    # ---- Assinaturas ----
    row[0] += 1
    r = row[0]
    merge(r, "A", "C")
    merge(r, "E", "G")
    ws[f"A{r}"].value, ws[f"A{r}"].alignment, ws[f"A{r}"].border = "Assinatura da Pessoa — Nome / Data", _C, Border(top=_THIN)
    ws[f"E{r}"].value, ws[f"E{r}"].alignment, ws[f"E{r}"].border = "Responsável pela Empresa — Nome / Data", _C, Border(top=_THIN)
    ws.row_dimensions[r].height = 34

    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = 9
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_area = f"A1:G{row[0]}"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
